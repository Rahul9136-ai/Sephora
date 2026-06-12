"""
ETL: aggregate raw Gladly contact-level exports (Ss.xlsx) into a clean
daily time series by Channel / Country / Language.

Channel coding
--------------
  EMAIL -> EM
  CHAT  -> CH
  VOICE -> VO

Country / Language coding
--------------------------
Voice rows already encode "<COUNTRY>-<LANGUAGE>" (e.g. US-EN, CA-FR, US-SP).
Email/Chat rows only encode country ("US", "CA-EN", "CA-FR", "ALL"); the
missing language is inferred (US -> EN, CA-EN -> EN, CA-FR -> FR). Rows with
country "ALL" (un-routed chat volume from before the Oct-2025 routing change)
are dropped - that bucket is discontinued and not part of any active series.

Each row is mapped to a Series_ID such as "VO US EN", "EM US EN",
"CH CA FR", following the convention requested for this project.

Output
------
data/processed/daily_contacts.csv with columns:
    Date, Channel, Country, Language, Series_ID, Contacts
"""

from __future__ import annotations

import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
import pandas as pd

SOURCE_PATH = Path(__file__).resolve().parents[2] / "data" / "raw" / "contacts_export.xlsx"
OUTPUT_PATH = Path(__file__).resolve().parents[2] / "data" / "processed" / "daily_contacts.csv"

CHANNEL_CODE = {
    "EMAIL": "EM",
    "CHAT": "CH",
    "VOICE": "VO",
}

# Email/Chat "Country" column does not always carry a language suffix.
# "ALL" (un-routed chat, discontinued) is intentionally absent - those rows
# are dropped in aggregate_email_chat.
EMAIL_CHAT_COUNTRY_MAP = {
    "US": ("US", "EN"),
    "CA-EN": ("CA", "EN"),
    "CA-FR": ("CA", "FR"),
}


def split_voice_country(raw: str) -> tuple[str, str]:
    """'US-EN' -> ('US', 'EN'); falls back gracefully for unexpected values."""
    if raw and "-" in raw:
        country, lang = raw.split("-", 1)
        return country, lang
    return (raw or "UNK"), "UNK"


def aggregate_email_chat(path: Path) -> dict[tuple, int]:
    """Aggregate the 'Email_Chat Export (Gladly)' sheet to (date, channel, country, lang) -> contacts."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["Email_Chat Export (Gladly)"]

    agg: dict[tuple, int] = defaultdict(int)
    for row in ws.iter_rows(min_row=2, values_only=True, max_col=13):
        date, channel_raw, country_raw, contacts = row[5], row[8], row[10], row[12]
        if date is None or channel_raw is None or contacts is None:
            continue

        channel = CHANNEL_CODE.get(channel_raw)
        if channel is None:
            continue

        mapped = EMAIL_CHAT_COUNTRY_MAP.get(country_raw)
        if mapped is None:
            continue
        country, lang = mapped
        agg[(pd.Timestamp(date).normalize(), channel, country, lang)] += contacts

    wb.close()
    return agg


def aggregate_voice(path: Path) -> dict[tuple, int]:
    """Aggregate the 'Voice Export (Gladly)' sheet to (date, channel, country, lang) -> contacts."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["Voice Export (Gladly)"]

    agg: dict[tuple, int] = defaultdict(int)
    for row in ws.iter_rows(min_row=2, values_only=True, max_col=8):
        date, channel_raw, country_raw, contacts = row[0], row[3], row[5], row[7]
        if date is None or channel_raw is None or contacts is None:
            continue

        channel = CHANNEL_CODE.get(channel_raw)
        if channel is None:
            continue

        country, lang = split_voice_country(country_raw)
        agg[(pd.Timestamp(date).normalize(), channel, country, lang)] += contacts

    wb.close()
    return agg


def build_daily_contacts(source_path: Path = SOURCE_PATH) -> pd.DataFrame:
    print(f"Reading {source_path} ...")
    email_chat_agg = aggregate_email_chat(source_path)
    print(f"  Email/Chat: {len(email_chat_agg):,} (date, series) cells")
    voice_agg = aggregate_voice(source_path)
    print(f"  Voice:      {len(voice_agg):,} (date, series) cells")

    combined: dict[tuple, int] = defaultdict(int)
    for agg in (email_chat_agg, voice_agg):
        for key, val in agg.items():
            combined[key] += val

    records = [
        {
            "Date": date,
            "Channel": channel,
            "Country": country,
            "Language": lang,
            "Series_ID": f"{channel} {country} {lang}",
            "Contacts": contacts,
        }
        for (date, channel, country, lang), contacts in combined.items()
    ]

    df = pd.DataFrame.from_records(records)
    df = df.sort_values(["Series_ID", "Date"]).reset_index(drop=True)
    return df


def main() -> None:
    df = build_daily_contacts()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(OUTPUT_PATH, index=False)
    print(f"\nWrote {len(df):,} rows to {OUTPUT_PATH}")

    print("\nSeries summary (total contacts, date range):")
    summary = (
        df.groupby("Series_ID")
        .agg(Total_Contacts=("Contacts", "sum"), Days=("Date", "nunique"),
             Start=("Date", "min"), End=("Date", "max"))
        .sort_values("Total_Contacts", ascending=False)
    )
    with pd.option_context("display.width", 160):
        print(summary)


if __name__ == "__main__":
    main()
